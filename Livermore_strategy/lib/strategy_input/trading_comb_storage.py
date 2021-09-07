# -*- encoding: utf-8 -*-
'''
@File    :   trading_comb_storage.py
@Time    :   2021/08/30 20:49:06
@Author  :   Wuhaotian 
@Version :   1.0
@Contact :   wuhaotian1007@gmail.com
'''

# here put the import lib
import openpyxl
import sqlite3
import configparser


class trading_comb_storage:
    def __init__(self, trading_comb_xlsx_path, trading_comb_db_path):
        # 读取全局配置
        def glob_conf():
            global onedrive_path
            conf = configparser.ConfigParser()
            '''读取配置文件'''
            # 全局文件路径
            conf.read('global.conf')
            # 获取指定section 的option值
            onedrive_path = conf.get("global", "Onedrive_path")

        glob_conf()

        '''建立数据库连接'''
        conn_trading_comb_db = sqlite3.connect(trading_comb_db_path)
        cur_trading_comb = conn_trading_comb_db.cursor()

        """ 新建数据库存储策略 """
        def create_database():
            # 新建table
            cur_trading_comb.execute('''CREATE TABLE IF NOT EXISTS trading_comb_list
                (trading_comb_ID      INTEGER        PRIMARY KEY,
                trading_comb_name     TEXT           NOT NULL,
                comb_category         TEXT           NOT NULL,
                trading_category_ID   INT          NOT NULL,
                note                  TEXT           NULL);''')

        create_database()

        """ 读取策略输入xlsx并入库 """
        def read_xlsx(xlsx_path):
            trading_comb_book = openpyxl.load_workbook(xlsx_path)
            sheet = trading_comb_book.active
            ''' 按行迭代策略信息 '''
            for row in range(2, sheet.max_row + 1):
                ''' 读取各列数据 '''
                trading_comb_ID = sheet.cell(row, 1).value
                trading_comb_name = sheet.cell(row, 2).value
                comb_category = sheet.cell(row, 3).value
                trading_category_ID = sheet.cell(row, 4).value
                note = sheet.cell(row, 5).value

                '''策略数据入库'''
                cur_trading_comb.execute("insert into trading_comb_list(trading_comb_ID, trading_comb_name, comb_category, trading_category_ID, note) \
                    VALUES(?, ?, ?, ?, ?)", (trading_comb_ID, trading_comb_name, comb_category, trading_category_ID, note))

        read_xlsx(onedrive_path + trading_comb_xlsx_path)

        # 上传并关闭数据库连接
        conn_trading_comb_db.commit()
        conn_trading_comb_db.close()
        print("策略基础-交易组合入库成功")


if __name__ == "__main__":
    trading_comb_storage(
        r"\work\strategy\strategy_input\base\交易组合.xlsx", r"database\trading_comb.db")
