# -*- encoding: utf-8 -*-
'''
@File    :   safty_strategy_storage.py
@Time    :   2021/08/30 23:03:22
@Author  :   Wuhaotian 
@Version :   1.0
@Contact :   wuhaotian1007@gmail.com
'''

# here put the import lib
import openpyxl
import sqlite3
import configparser


class safty_strategy_storage:
    def __init__(self, safty_strategy_xlsx_path, safty_strategy_db_path):
        # 读取全局配置
        def glob_conf():
            global onedrive_path
            conf = configparser.ConfigParser()
            '''读取配置文件'''
            # 全局文件路径
            conf.read('global.conf')
            # 获取指定section 的option值
            onedrive_path = conf.get("global", "Onedrive_path")

        glob_conf()

        '''建立数据库连接'''
        conn_safty_strategy_db = sqlite3.connect(
            safty_strategy_db_path)
        cur_safty_strategy = conn_safty_strategy_db.cursor()

        """ 新建数据库存储策略 """
        def create_database():
            # 新建table
            cur_safty_strategy.execute('''CREATE TABLE IF NOT EXISTS safty_strategy_list
                (safty_strategy_ID       INTEGER        PRIMARY KEY,
                soft_limit               INT           NOT NULL,
                safty_line               INT           NOT NULL,
                forced_closing_line      FLOAT         NOT NULL,
                interest_withdraw_ratio  TEXT          NOT NULL,
                withdraw_ratio_step      FLOAT          NOT NULL,
                note                     TEXT          NULL);''')

        create_database()

        """ 读取策略输入xlsx并入库 """
        def read_xlsx(xlsx_path):
            safty_strategy_book = openpyxl.load_workbook(xlsx_path)
            sheet = safty_strategy_book.active
            ''' 按行迭代策略信息 '''
            for row in range(2, sheet.max_row + 1):
                ''' 读取各列数据 '''
                safty_strategy_ID = sheet.cell(row, 1).value
                soft_limit = sheet.cell(row, 2).value
                safty_line = sheet.cell(row, 3).value
                forced_closing_line = sheet.cell(row, 4).value
                interest_withdraw_ratio = sheet.cell(row, 5).value
                withdraw_ratio_step = sheet.cell(row, 6).value
                note = sheet.cell(row, 7).value

                '''策略数据入库'''
                cur_safty_strategy.execute("insert into safty_strategy_list(safty_strategy_ID, soft_limit, safty_line, forced_closing_line, interest_withdraw_ratio, withdraw_ratio_step, note) \
                    VALUES(?, ?, ?, ?, ?, ?, ?)", (safty_strategy_ID, soft_limit, safty_line, forced_closing_line, interest_withdraw_ratio, withdraw_ratio_step, note))

        read_xlsx(onedrive_path + safty_strategy_xlsx_path)

        # 上传并关闭数据库连接
        conn_safty_strategy_db.commit()
        conn_safty_strategy_db.close()
        print("策略基础-安全策略入库成功")


if __name__ == "__main__":
    safty_strategy_storage(
        r"\work\strategy\strategy_input\base\安全策略.xlsx", r"database\safty_strategy.db")
