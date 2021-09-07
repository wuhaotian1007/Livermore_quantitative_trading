# -*- encoding: utf-8 -*-
'''
@File    :   buy_strategy_storage.py
@Time    :   2021/08/30 23:03:22
@Author  :   Wuhaotian 
@Version :   1.0
@Contact :   wuhaotian1007@gmail.com
'''

# here put the import lib
import openpyxl
import sqlite3
import configparser


class buy_strategy_storage:
    def __init__(self, buy_strategy_xlsx_path, buy_strategy_db_path):
        # 读取全局配置
        def glob_conf():
            global onedrive_path
            conf = configparser.ConfigParser()
            '''读取配置文件'''
            # 全局文件路径
            conf.read('global.conf')
            # 获取指定section 的option值
            onedrive_path = conf.get("global", "Onedrive_path")

        glob_conf()

        '''建立数据库连接'''
        conn_buy_strategy_db = sqlite3.connect(
            buy_strategy_db_path)
        cur_buy_strategy = conn_buy_strategy_db.cursor()

        """ 新建数据库存储策略 """
        def create_database():
            # 新建table
            cur_buy_strategy.execute('''CREATE TABLE IF NOT EXISTS buy_strategy_list
                (buy_strategy_ID         INTEGER        PRIMARY KEY,
                position_rate            TEXT           NOT NULL,
                position_rate_step       TEXT           NOT NULL,
                step_method              TEXT           NOT NULL,
                per_step_ratio           TEXT           NOT NULL,
                increment_step           INT            NULL,
                buy_mode                 TEXT           NOT NULL,
                note                     TEXT           NULL);''')

        create_database()

        """ 读取策略输入xlsx并入库 """
        def read_xlsx(xlsx_path):
            buy_strategy_book = openpyxl.load_workbook(xlsx_path)
            sheet = buy_strategy_book.active
            ''' 按行迭代策略信息 '''
            for row in range(2, sheet.max_row + 1):
                ''' 读取各列数据 '''
                buy_strategy_ID = sheet.cell(row, 1).value
                position_rate = sheet.cell(row, 2).value
                position_rate_step = sheet.cell(row, 3).value
                step_method = sheet.cell(row, 4).value
                per_step_ratio = sheet.cell(row, 5).value
                increment_step = sheet.cell(row, 6).value
                buy_mode = sheet.cell(row, 7).value
                note = sheet.cell(row, 8).value

                '''策略数据入库'''
                cur_buy_strategy.execute("insert into buy_strategy_list(buy_strategy_ID, position_rate, position_rate_step, step_method, per_step_ratio, increment_step, buy_mode, note) \
                    VALUES(?, ?, ?, ?, ?, ?, ?, ?)", (buy_strategy_ID, position_rate, position_rate_step, step_method, per_step_ratio, increment_step, buy_mode, note))

        read_xlsx(onedrive_path + buy_strategy_xlsx_path)

        # 上传并关闭数据库连接
        conn_buy_strategy_db.commit()
        conn_buy_strategy_db.close()
        print("策略基础-买入策略入库成功")


if __name__ == "__main__":
    buy_strategy_storage(
        r"\work\strategy\strategy_input\base\买入策略.xlsx", r"database\buy_strategy.db")
