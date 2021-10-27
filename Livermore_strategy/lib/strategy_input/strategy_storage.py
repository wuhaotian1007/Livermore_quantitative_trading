# -*- encoding: utf-8 -*-
'''
@File    :   strategy_storage.py
@Time    :   2021/08/29 16:40:21
@Author  :   Wuhaotian 
@Version :   1.0
@Contact :   wuhaotian1007@gmail.com
'''

# here put the import lib
import openpyxl
import sqlite3
import configparser


class strategy_storage:
    def __init__(self, strategy_xlsx_path, strategy_db_path):
        # 读取全局配置
        def glob_conf():
            global onedrive_path
            conf = configparser.ConfigParser()
            '''读取配置文件'''
            # 全局文件路径
            conf.read('global.conf')
            # 获取指定section 的option值
            onedrive_path = conf.get("global", "Onedrive_path")

        glob_conf()

        '''建立数据库连接'''
        conn_strategy_db = sqlite3.connect(strategy_db_path)
        cur_strategy = conn_strategy_db.cursor()

        """ 新建数据库存储策略 """
        def create_database():
            # 新建table
            cur_strategy.execute('''CREATE TABLE IF NOT EXISTS strategy_list
                (strategy_ID          INTEGER        PRIMARY KEY,
                time_interval         TEXT           NOT NULL,
                trading_comb_ID       INT            NOT NULL,
                initial_surplus       FLOAT          NOT NULL,
                safty_strategy_ID     INT            NOT NULL,
                signal_comb_ID_list   TEXT           NOT NULL,
                trading_frequency     TEXT           NOT NULL,
                buy_strategy_ID       INT            NOT NULL,
                note                  TEXT           NULL);''')

        create_database()

        """ 读取策略输入xlsx并入库 """
        def read_xlsx(xlsx_path):
            strategy_book = openpyxl.load_workbook(xlsx_path)
            sheet = strategy_book.active
            ''' 按行迭代策略信息 '''
            for row in range(2, sheet.max_row + 1):
                ''' 读取各列数据 '''
                strategy_ID = sheet.cell(row, 1).value
                time_interval = sheet.cell(row, 2).value
                trading_comb_ID = sheet.cell(row, 3).value
                initial_surplus = sheet.cell(row, 4).value
                safty_strategy_ID = sheet.cell(row, 5).value
                signal_comb_ID_list = sheet.cell(row, 6).value
                trading_frequency = sheet.cell(row, 7).value
                buy_strategy_ID = sheet.cell(row, 8).value
                note = sheet.cell(row, 9).value

                '''策略数据入库'''
                cur_strategy.execute("insert into strategy_list(strategy_ID, time_interval, trading_comb_ID, initial_surplus, safty_strategy_ID, signal_comb_ID_list, trading_frequency, buy_strategy_ID, note) \
                    VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)", (strategy_ID, time_interval, trading_comb_ID, initial_surplus, safty_strategy_ID, signal_comb_ID_list, trading_frequency, buy_strategy_ID, note))

        read_xlsx(onedrive_path + strategy_xlsx_path)

        # 上传并关闭数据库连接
        conn_strategy_db.commit()
        conn_strategy_db.close()
        print("策略入库成功")


if __name__ == "__main__":
    strategy_storage(r"\work\strategy\strategy_input\交易策略.xlsx",
                     r"database\strategy.db")
