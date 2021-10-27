# -*- encoding: utf-8 -*-
'''
@File    :   trading_siganl_comb_storage.py
@Time    :   2021/08/30 22:38:48
@Author  :   Wuhaotian 
@Version :   1.0
@Contact :   wuhaotian1007@gmail.com
'''

# here put the import lib
import openpyxl
import sqlite3
import configparser


class trading_signal_comb_storage:
    def __init__(self, trading_signal_comb_xlsx_path, trading_signal_comb_db_path):
        # 读取全局配置
        def glob_conf():
            global onedrive_path
            conf = configparser.ConfigParser()
            '''读取配置文件'''
            # 全局文件路径
            conf.read('global.conf')
            # 获取指定section 的option值
            onedrive_path = conf.get("global", "Onedrive_path")

        glob_conf()

        '''建立数据库连接'''
        conn_trading_signal_comb_db = sqlite3.connect(
            trading_signal_comb_db_path)
        cur_trading_signal_comb = conn_trading_signal_comb_db.cursor()

        """ 新建数据库存储策略 """
        def create_database():
            # 新建table
            cur_trading_signal_comb.execute('''CREATE TABLE IF NOT EXISTS trading_signal_comb_list
                (trading_signal_comb_ID  INTEGER        PRIMARY KEY,
                trading_direction        TEXT           NOT NULL,
                buy_signal               TEXT           NOT NULL,
                close_signal             TEXT           NOT NULL,
                signal_blacklist         TEXT           NOT NULL,
                note                     TEXT           NULL);''')

        create_database()

        """ 读取策略输入xlsx并入库 """
        def read_xlsx(xlsx_path):
            trading_signal_comb_book = openpyxl.load_workbook(xlsx_path)
            sheet = trading_signal_comb_book.active
            ''' 按行迭代策略信息 '''
            for row in range(2, sheet.max_row + 1):
                ''' 读取各列数据 '''
                trading_signal_comb_ID = sheet.cell(row, 1).value
                trading_direction = sheet.cell(row, 2).value
                buy_signal = sheet.cell(row, 3).value
                close_signal = sheet.cell(row, 4).value
                signal_blacklist = sheet.cell(row, 5).value
                note = sheet.cell(row, 6).value

                '''策略数据入库'''
                cur_trading_signal_comb.execute("insert into trading_signal_comb_list(trading_signal_comb_ID, trading_direction, buy_signal, close_signal, signal_blacklist, note) \
                    VALUES(?, ?, ?, ?, ?, ?)", (trading_signal_comb_ID, trading_direction, buy_signal, close_signal, signal_blacklist, note))

        read_xlsx(onedrive_path + trading_signal_comb_xlsx_path)

        # 上传并关闭数据库连接
        conn_trading_signal_comb_db.commit()
        conn_trading_signal_comb_db.close()
        print("策略基础-交易信号组合入库成功")

if __name__ == "__main__":
    trading_signal_comb_storage(
        r"\work\strategy\strategy_input\base\交易信号组合.xlsx", r"database\trading_signal_comb.db")
