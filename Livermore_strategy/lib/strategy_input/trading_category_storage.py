# -*- encoding: utf-8 -*-
'''
@File    :   trading_category_storage.py
@Time    :   2021/09/02 16:56:13
@Author  :   Wuhaotian 
@Version :   1.0
@Contact :   wuhaotian1007@gmail.com
'''

# here put the import lib
# here put the import lib
import openpyxl
import sqlite3
import configparser


class trading_category_storage:
    def __init__(self, trading_category_xlsx_path, trading_category_db_path):
        # 读取全局配置
        def glob_conf():
            global onedrive_path
            conf = configparser.ConfigParser()
            '''读取配置文件'''
            # 全局文件路径
            conf.read('global.conf')
            # 获取指定section 的option值
            onedrive_path = conf.get("global", "Onedrive_path")

        glob_conf()

        '''建立数据库连接'''
        conn_trading_category_db = sqlite3.connect(trading_category_db_path)
        cur_trading_category = conn_trading_category_db.cursor()

        """ 新建数据库存储策略 """
        def create_database():
            # 新建table
            cur_trading_category.execute('''CREATE TABLE IF NOT EXISTS trading_category_list
                (trading_category_ID    INTEGER        PRIMARY KEY,
                category_name           TEXT           NOT NULL,
                volume_per_hand         TEXT           NOT NULL,
                exact_digit             INT            NOT NULL,
                three_points            FLOAT          NOT NULL,
                note                    TEXT           NULL);''')

        create_database()

        """ 读取策略输入xlsx并入库 """
        def read_xlsx(xlsx_path):
            strategy_book = openpyxl.load_workbook(xlsx_path)
            sheet = strategy_book.active
            ''' 按行迭代策略信息 '''
            a = sheet.max_row
            for row in range(2, sheet.max_row + 1):
                ''' 读取各列数据 '''
                trading_category_ID = sheet.cell(row, 1).value
                category_name = sheet.cell(row, 2).value
                volume_per_hand = sheet.cell(row, 3).value
                exact_digit = sheet.cell(row, 4).value
                three_points = sheet.cell(row, 5).value
                note = sheet.cell(row, 6).value

                '''策略数据入库'''
                cur_trading_category.execute("insert into trading_category_list(trading_category_ID, category_name, volume_per_hand, exact_digit, three_points, note) \
                    VALUES(?, ?, ?, ?, ?, ?)", (trading_category_ID, category_name, volume_per_hand, exact_digit, three_points, note))

        read_xlsx(onedrive_path + trading_category_xlsx_path)

        # 上传并关闭数据库连接
        conn_trading_category_db.commit()
        conn_trading_category_db.close()
        print("策略基础-交易品类入库成功")


if __name__ == "__main__":
    trading_category_storage(r"\work\strategy\strategy_input\base\交易品类.xlsx",
                     r"database\trading_category.db")